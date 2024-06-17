from flask import Flask, request, render_template, redirect, url_for
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/record', methods=['POST'])
def record():
    try:
        fitness = request.form['fitness']
        comments = request.form.get('comments', 'N/A') or 'N/A'
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if fitness == 'no':
            # If not about fitness, only write the timestamp, fitness response, and comments
            data = [timestamp, fitness, 'N/A', 'N/A', 'N/A', 'N/A', comments]
        else:
            steroids_video = request.form.get('steroidsVideo', 'N/A') or 'N/A'
            steroids_caption = request.form.get('steroidsCaption', 'N/A') or 'N/A'
            mention_sentiment = request.form.get('mentionSentiment', 'N/A') or 'N/A'
            sale_steroids = request.form.get('saleSteroids', 'N/A') or 'N/A'
            data = [timestamp, fitness, steroids_video, steroids_caption, mention_sentiment, sale_steroids, comments]

        # Ensure the file exists before trying to open it
        if not os.path.exists('data.xlsx'):
            # Create the workbook and add headers if it doesn't exist
            from openpyxl import Workbook
            workbook = Workbook()
            sheet = workbook.active
            headers = ['Timestamp', 'Fitness', 'Steroids Video', 'Steroids Caption', 'Mention Sentiment', 'Sale of Steroids', 'Comments']
            sheet.append(headers)
            workbook.save('data.xlsx')

        # Load the workbook and select the active worksheet
        workbook = load_workbook('data.xlsx')
        sheet = workbook.active

        # Find the next available row
        next_row = sheet.max_row + 1

        # Append the data
        for col_num, value in enumerate(data, start=1):
            sheet.cell(row=next_row, column=col_num).value = value

        workbook.save('data.xlsx')

        return redirect(url_for('index'))
    except Exception as e:
        return f"An error occurred: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True)
